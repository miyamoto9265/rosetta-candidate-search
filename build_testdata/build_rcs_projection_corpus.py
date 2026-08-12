#!/usr/bin/env python3
"""Build RCS evaluation corpus from wholebif_extract.xlsx projections sheet.

Extracts sender/receiver region names (plus fullname) from mammalian
projection records, drops UNKNOWN and cortical-layer-only labels, and
writes a unique-name CSV compatible with rcs/rcs_test_list.py
(first column = query).

Example (from repo root):
    python build_testdata/build_rcs_projection_corpus.py
    python build_testdata/build_rcs_projection_corpus.py --xlsx wholebif_extract.xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path

ROOT = Path(__file__).resolve().parent
REPO_ROOT = ROOT.parent
DEFAULT_XLSX = REPO_ROOT / "wholebif_extract.xlsx"
DEFAULT_OUTPUT = ROOT / "rcs_projection_corpus.csv"

OUTPUT_FIELDS = (
    "structure_name",
    "fullname",
    "species",
    "paper",
    "n_mentions",
    "n_papers",
)

# Cortical / laminar labels that are not region names for RCS.
LAYER_ONLY_RE = re.compile(
    r"^L(?:[1-6]|I{1,3}|IV|V|VI)[A-Za-z]?$",
    re.IGNORECASE,
)

UNKNOWN_NAMES = {
    "",
    "<unknown>",
    "unknown",
    "n/a",
    "na",
    "none",
    "null",
}

# Substrings that mark non-mammal rows (matched against taxon / taxon_common).
NON_MAMMAL_SUBSTRINGS = (
    "drosophila",
    "danio",
    "zebrafish",
    "zebra finch",
    "bird",
    "chicken",
    "chick",
    "columba",
    "pigeon",
    "crow",
    "octopus",
    "locust",
    "neomysis",
    "alligator",
    "cichlid",
    "goby",
    "salmon",
    "xenopus",
    "tadpole",
    "frog",
    "fish",
    "fly (",
    "insect",
    "cephalopod",
)

# Positive mammal hints for free-text taxon labels.
MAMMAL_SUBSTRINGS = (
    "mouse",
    "mus ",
    "rat",
    "rattus",
    "human",
    "homo ",
    "macaque",
    "macaca",
    "monkey",
    "primate",
    "marmoset",
    "callithrix",
    "cat",
    "felis",
    "ferret",
    "mustela",
    "rabbit",
    "oryctolagus",
    "rodent",
    "shrew",
    "gerbil",
    "vole",
    "pig",
    "sheep",
    "galago",
    "baboon",
    "chimpanzee",
    "ape",
    "canine",
    "dog",
    "hamster",
    "guinea",
    "squirrel monkey",
    "acomys",
)

# Canonical species labels for RCS corpus convention.
SPECIES_CANONICAL = {
    "mus musculus": "Mouse",
    "mouse": "Mouse",
    "rattus norvegicus": "Rat",
    "rat": "Rat",
    "homo sapiens": "Human",
    "human": "Human",
    "human (graft); hie-injured brain host": "Human",
    "macaca mulatta": "Macaque",
    "macaque": "Macaque",
    "rhesus macaque": "Macaque",
    "monkey": "Macaque",
    "primate": "Primate",
    "callithrix jacchus": "Marmoset",
    "common marmoset": "Marmoset",
    "marmoset": "Marmoset",
    "felis catus": "Cat",
    "cat": "Cat",
    "mustela putorius furo": "Ferret",
    "ferret": "Ferret",
    "oryctolagus cuniculus": "Rabbit",
    "rabbit": "Rabbit",
    "rodent": "Rodent",
    "rodents": "Rodent",
    "tree shrew": "Tree shrew",
    "gerbil": "Gerbil",
    "prairie vole": "Prairie vole",
    "pig": "Pig",
    "minipig": "Pig",
    "sheep": "Sheep",
    "galago": "Galago",
    "acomys cahirinus": "Spiny mouse",
    "squirrel monkey": "Squirrel monkey",
    "mongolian gerbil": "Gerbil",
    "common marmoset monkey": "Marmoset",
    "cebus monkey": "Cebus monkey",
    "nonhuman primates": "Primate",
    "non-human primate": "Primate",
    "syrian hamster": "Hamster",
    "macaca fascicularis, macaca mulatta": "Macaque",
}


@dataclass
class NameAgg:
    display_counts: Counter[str] = field(default_factory=Counter)
    fullname_counts: Counter[str] = field(default_factory=Counter)
    species_counts: Counter[str] = field(default_factory=Counter)
    paper_counts: Counter[str] = field(default_factory=Counter)
    n_mentions: int = 0

    def add(self, display: str, fullname: str, species: str, paper: str) -> None:
        self.display_counts[display] += 1
        if fullname and fullname.casefold() != display.casefold():
            self.fullname_counts[fullname] += 1
        elif fullname:
            # Same as short name; still record so majority can surface it.
            self.fullname_counts[fullname] += 1
        if species:
            self.species_counts[species] += 1
        if paper:
            self.paper_counts[paper] += 1
        self.n_mentions += 1

    def structure_name(self) -> str:
        return self.display_counts.most_common(1)[0][0]

    def fullname(self) -> str:
        if not self.fullname_counts:
            return ""
        return self.fullname_counts.most_common(1)[0][0]

    def species(self) -> str:
        if not self.species_counts:
            return ""
        return self.species_counts.most_common(1)[0][0]

    def paper(self) -> str:
        if not self.paper_counts:
            return ""
        return self.paper_counts.most_common(1)[0][0]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build RCS projection-region corpus from wholebif_extract.xlsx."
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"Input workbook (default: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output CSV (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--sheet",
        default="projections",
        help="Worksheet name (default: projections)",
    )
    return parser.parse_args()


def is_layer_only(name: str) -> bool:
    return bool(LAYER_ONLY_RE.fullmatch(name.strip()))


def is_unknown(name: str) -> bool:
    return name.strip().casefold() in UNKNOWN_NAMES


def _taxon_blob(taxon: str, taxon_common: str) -> str:
    return f"{taxon} {taxon_common}".strip().casefold()


def is_mammal_row(taxon: str, taxon_common: str) -> bool:
    """Keep explicit mammals and unlabeled rows; drop clear non-mammals."""
    blob = _taxon_blob(taxon, taxon_common)
    if not blob:
        return True
    if any(marker in blob for marker in NON_MAMMAL_SUBSTRINGS):
        return False
    for value in (taxon_common, taxon):
        if value.strip().casefold() in SPECIES_CANONICAL:
            return True
    if any(marker in blob for marker in MAMMAL_SUBSTRINGS):
        return True
    # Labeled but unrecognized: exclude to keep the corpus mammal-centered.
    return False


def canonicalize_species(taxon: str, taxon_common: str) -> str:
    for value in (taxon_common, taxon):
        key = value.strip().casefold()
        if key in SPECIES_CANONICAL:
            return SPECIES_CANONICAL[key]
    if taxon_common.strip():
        return taxon_common.strip().title()
    if taxon.strip():
        return taxon.strip()
    return ""


def load_projection_rows(xlsx: Path, sheet: str) -> list[dict[str, str]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required. Install with: python -m pip install openpyxl"
        ) from exc

    if not xlsx.is_file():
        raise SystemExit(f"Input workbook not found: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")

    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    headers = next(it, None)
    if not headers:
        wb.close()
        raise SystemExit(f"Sheet {sheet!r} is empty")

    hmap = {str(h): i for i, h in enumerate(headers) if h is not None}
    required = (
        "sender",
        "receiver",
        "sender_fullname",
        "receiver_fullname",
        "paper_id",
        "taxon",
        "taxon_common",
    )
    missing = [c for c in required if c not in hmap]
    if missing:
        wb.close()
        raise SystemExit(f"Missing columns: {missing}")

    rows: list[dict[str, str]] = []
    for raw in it:
        def cell(col: str) -> str:
            val = raw[hmap[col]]
            return "" if val is None else str(val).strip()

        rows.append(
            {
                "sender": cell("sender"),
                "receiver": cell("receiver"),
                "sender_fullname": cell("sender_fullname"),
                "receiver_fullname": cell("receiver_fullname"),
                "paper_id": cell("paper_id"),
                "taxon": cell("taxon"),
                "taxon_common": cell("taxon_common"),
            }
        )
    wb.close()
    return rows


def build_corpus(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    aggs: dict[str, NameAgg] = defaultdict(NameAgg)
    skipped_non_mammal = 0
    skipped_unknown = 0
    skipped_layer = 0
    kept_mentions = 0

    for row in rows:
        if not is_mammal_row(row["taxon"], row["taxon_common"]):
            skipped_non_mammal += 1
            continue

        species = canonicalize_species(row["taxon"], row["taxon_common"])
        paper = row["paper_id"]

        for name, fullname in (
            (row["sender"], row["sender_fullname"]),
            (row["receiver"], row["receiver_fullname"]),
        ):
            if is_unknown(name):
                skipped_unknown += 1
                continue
            if is_layer_only(name):
                skipped_layer += 1
                continue

            key = name.casefold()
            aggs[key].add(
                display=name,
                fullname=fullname,
                species=species,
                paper=paper,
            )
            kept_mentions += 1

    out: list[dict[str, str]] = []
    for agg in aggs.values():
        out.append(
            {
                "structure_name": agg.structure_name(),
                "fullname": agg.fullname(),
                "species": agg.species(),
                "paper": agg.paper(),
                "n_mentions": str(agg.n_mentions),
                "n_papers": str(len(agg.paper_counts)),
            }
        )

    out.sort(key=lambda r: (-int(r["n_mentions"]), r["structure_name"].casefold()))

    print(
        f"rows_in={len(rows)} kept_mentions={kept_mentions} "
        f"unique_names={len(out)} skipped_non_mammal_rows={skipped_non_mammal} "
        f"skipped_unknown={skipped_unknown} skipped_layer={skipped_layer}",
        file=sys.stderr,
    )
    return out


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    rows = load_projection_rows(args.xlsx, args.sheet)
    corpus = build_corpus(rows)
    write_csv(args.output, corpus)
    print(f"Wrote {len(corpus)} rows -> {args.output}")


if __name__ == "__main__":
    main()
